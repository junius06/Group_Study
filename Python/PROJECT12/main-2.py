# pandas 라이브러리로 값을 엑셀로 저장 후 출력

from openpyxl import load_workbook

# 엑셀 파일을 읽어온다.
load_wb = load_workbook(r'12. 엑셀의 정보를 불러와 수료증 자동 생성\수료증명단.xlsx')
# 읽어온 엑셀 파일에서 활성화된 시트를 불러온다. >> 활성화된 시트가 여러개일 경우?
load_ws = load_wb.active

name_list = []
birthday_list = []
ho_list = []
for i in range(1,load_ws.max_row + 1):
    name_list.append(load_ws.cell(i, 1).value)
    birthday_list.append(load_ws.cell(i, 2).value)
    ho_list.append(load_ws.cell(i, 3).value)

print(name_list)
print(birthday_list)
print(ho_list)