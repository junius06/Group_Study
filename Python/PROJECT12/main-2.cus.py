from openpyxl import load_workbook

def read_excel(file_path, sheet_name):
    # 엑셀 파일을 읽어온다.
    load_wb = load_workbook(file_path)
    # 시트명에 해당하는 시트를 선택한다.
    load_ws = load_wb[sheet_name]

    name_list = []
    birthday_list = []
    ho_list = []

    for i in range(1,load_ws.max_row + 1):
        name_list.append(load_ws.cell(i, 1).value)
        birthday_list.append(load_ws.cell(i, 2).value)
        ho_list.append(load_ws.cell(i, 3).value)

    return name_list, birthday_list, ho_list

file_path = input("엑셀 파일 경로 : ")
sheet_name = input("엑셀 시트 이름 : ")

name_list, birthday_list, ho_list = read_excel(file_path, sheet_name)

print(name_list)
print(birthday_list)
print(ho_list)
