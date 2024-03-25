from openpyxl import load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

load_wb = load_workbook(r"email_list.xlsx", data_only=True)
load_ws = load_wb.active

for i in range(1, load_ws.max_row + 1):
    recv_email_value = load_ws.cell(i, 1).value
    print("Success: ", recv_email_value)
    try:
        send_email = "joohee5079@naver.com"
        send_pwd = "password"
        
        recv_email = recv_email_value
        
        msg = MIMEMultipart()
        
        msg['Subject'] = "자동발송 메일 테스트"
        msg['From'] = send_email
        msg['To'] = recv_email
        
        text = """
                테스트 이메일
                """
        
        msg.attach(MIMEText(text))
        
        mailServer = smtplib.SMTP_SSL('smtp.naver.com')
        mailServer.login(send_email, send_pwd)
        mailServer.sendmail(send_email, recv_email, msg.as_string())
        mailServer.quit()
    except:
        print("Error: ", recv_email_value)