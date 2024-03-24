import requests # http API 요청
import re # 정규식 표현 사용
from openpyxl import load_workbook # 기존의 엑셀파일 로드
from openpyxl import Workbook # 새로운 엑셀파일 생성

url = 'https://news.v.daum.net/v/20211129144552297'

headers = {
    'User-Agent' : 'Mozilla/5.0',
    'Content-Type' : 'text/html, charset=utf-8'
    }

response = requests.get(url, headers=headers)

results = list(set(re.findall(r'[\w\.-]+@[\w\.-]+', response.text)))

print(results)

try: # 이메일 주소를 기록할 엑셀 파일을 로드하고, 
    wb = load_workbook(r"13. 이메일을 수집하여 엑셀에 기록하기\email.xlsx", data_only=True)
    sheet = wb.active
except: # 파일이 존재하지 않으면 새로운 Workbook을 생성한다.
    wb = Workbook()
    sheet = wb.active

for result in results: # 추출된 이메일 주소를 반복하면서 중복은 제거하고 엑셀 시트에 저장한다.
    sheet.append([result])

# 변경된 엑셀파일을 저장한다. 절대경로를 지정하지 않으면 코드가 실행되는 위치에 상대경로로 저장된다.
wb.save(r"/Users/jhkim/learning/my-github/Group_Study/Python/PROJECT13/13. 이메일을 수집하여 엑셀에 기록하기\email.xlsx")